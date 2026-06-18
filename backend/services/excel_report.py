"""
테스트 수행 내역서 Excel 생성 (openpyxl) — Discord 알림 첨부용
"""

import io
from datetime import datetime, timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

_GREEN  = "FF22C55E"
_RED    = "FFEF4444"
_ORANGE = "FFF97316"
_GRAY   = "FF6B7280"
_HEADER = "FF1E293B"
_WHITE  = "FFFFFFFF"
_LIGHT  = "FFF1F5F9"


def _thin():
    s = Side(style="thin", color="FFD1D5DB")
    return Border(left=s, right=s, top=s, bottom=s)


def build_excel_report(payload: dict) -> bytes:
    """payload 기반으로 테스트 수행 내역서 .xlsx 를 bytes로 반환"""
    run_id       = payload.get("run_id", 0)
    label        = payload.get("label") or f"Run #{run_id}"
    project      = payload.get("project_name", "—")
    base_url     = payload.get("base_url", "—")
    total        = payload.get("total", 0)
    fail         = payload.get("fail", 0)
    passed       = total - fail
    now          = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    case_results = payload.get("case_results", [])
    mgr_snapshot = payload.get("mgr_snapshot", [])

    # case_id → mgr 정보 맵
    mgr_map = {c["id"]: c for c in mgr_snapshot}

    wb = Workbook()
    ws = wb.active
    ws.title = "테스트 수행 내역서"

    # ── 헤더 정보 섹션 ──
    ws.merge_cells("A1:G1")
    ws["A1"] = "테스트 수행 내역서"
    ws["A1"].font = Font(bold=True, size=14, color=_WHITE)
    ws["A1"].fill = PatternFill("solid", fgColor=_HEADER)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    meta = [
        ("프로젝트", project),
        ("실행 레이블", label),
        ("실행 ID", f"#{run_id}"),
        ("일시", now),
        ("대상 서버", base_url),
        ("전체", f"{total}건"),
        ("통과", f"{passed}건"),
        ("실패", f"{fail}건"),
    ]
    for i, (k, v) in enumerate(meta, start=2):
        ws.merge_cells(f"A{i}:B{i}")
        ws[f"A{i}"] = k
        ws[f"A{i}"].font = Font(bold=True, size=10)
        ws[f"A{i}"].fill = PatternFill("solid", fgColor=_LIGHT)
        ws[f"A{i}"].alignment = Alignment(vertical="center")
        ws.merge_cells(f"C{i}:G{i}")
        ws[f"C{i}"] = v
        ws[f"C{i}"].font = Font(size=10)
        ws[f"C{i}"].alignment = Alignment(vertical="center")
        for col in "ABCDEFG":
            ws[f"{col}{i}"].border = _thin()
        ws.row_dimensions[i].height = 18

    # ── 케이스 결과 테이블 ──
    table_start = len(meta) + 3

    col_headers = ["번호", "케이스 ID", "케이스명", "메서드", "엔드포인트", "카테고리", "결과"]
    col_widths   = [6,      14,          30,          9,        40,            14,         8]

    for col_idx, (header, width) in enumerate(zip(col_headers, col_widths), start=1):
        cell = ws.cell(row=table_start, column=col_idx, value=header)
        cell.font = Font(bold=True, size=10, color=_WHITE)
        cell.fill = PatternFill("solid", fgColor=_HEADER)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _thin()
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[table_start].height = 20

    for row_offset, cr in enumerate(case_results, start=1):
        row = table_start + row_offset
        case_id = cr.get("case_id", "")
        mgr     = mgr_map.get(case_id, {})
        pf      = cr.get("pf", "")

        values = [
            row_offset,
            case_id,
            mgr.get("name", ""),
            mgr.get("method", ""),
            mgr.get("endpoint", ""),
            mgr.get("catId", ""),
            pf,
        ]
        bg = _LIGHT if row_offset % 2 == 0 else _WHITE
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.font = Font(size=10)
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.border = _thin()
            cell.alignment = Alignment(vertical="center")
            if col_idx == 7:  # 결과 컬럼
                if pf == "Pass":
                    cell.font = Font(size=10, bold=True, color=_GREEN)
                elif pf == "Fail":
                    cell.font = Font(size=10, bold=True, color=_RED)
        ws.row_dimensions[row].height = 16

    # 빈 케이스 결과 시
    if not case_results:
        no_data_row = table_start + 1
        ws.merge_cells(f"A{no_data_row}:G{no_data_row}")
        ws[f"A{no_data_row}"] = "케이스 결과 없음"
        ws[f"A{no_data_row}"].alignment = Alignment(horizontal="center")
        ws[f"A{no_data_row}"].font = Font(size=10, color=_GRAY)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
