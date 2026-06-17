from fastapi import APIRouter, UploadFile, File, HTTPException, Form
from pydantic import BaseModel
from typing import Optional
import openpyxl
import io
import re

router = APIRouter(prefix="/api/excel", tags=["excel"])


class ColumnMapping(BaseModel):
    id_col: Optional[str] = None        # 케이스 ID 컬럼명
    name_col: str = "테스트 항목명"       # 테스트 항목명 컬럼명
    type_col: Optional[str] = None      # Positive/Negative 구분
    input_col: Optional[str] = None     # 입력값
    expected_col: Optional[str] = None  # 기대 결과
    category_col: Optional[str] = None  # 카테고리
    sheet_name: Optional[str] = None    # 시트명 (없으면 첫 번째 시트)
    header_row: int = 1                 # 헤더가 있는 행 번호 (1-based)


@router.post("/parse")
async def parse_excel(
    file: UploadFile = File(...),
    name_col: str = Form("테스트 항목명"),
    id_col: str = Form(""),
    type_col: str = Form(""),
    input_col: str = Form(""),
    expected_col: str = Form(""),
    category_col: str = Form(""),
    sheet_name: str = Form(""),
    header_row: int = Form(1),
):
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="xlsx 또는 xls 파일만 지원합니다")

    contents = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"엑셀 파일 읽기 실패: {str(e)}")

    # 시트 선택
    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < header_row:
        raise HTTPException(status_code=400, detail="헤더 행을 찾을 수 없습니다")

    # 헤더 파싱 (0-based index로 변환)
    header = [str(h).strip() if h else "" for h in rows[header_row - 1]]

    def col_idx(col_name: str) -> Optional[int]:
        if not col_name:
            return None
        col_name = col_name.strip()
        for i, h in enumerate(header):
            if h == col_name or col_name.lower() in h.lower():
                return i
        return None

    idx_name = col_idx(name_col)
    idx_id = col_idx(id_col) if id_col else None
    idx_type = col_idx(type_col) if type_col else None
    idx_input = col_idx(input_col) if input_col else None
    idx_expected = col_idx(expected_col) if expected_col else None
    idx_category = col_idx(category_col) if category_col else None

    if idx_name is None:
        raise HTTPException(
            status_code=400,
            detail=f"'{name_col}' 컬럼을 찾을 수 없습니다. 실제 헤더: {header}"
        )

    cases = []
    categories = set()
    auto_id = 1

    for row in rows[header_row:]:
        name = str(row[idx_name]).strip() if row[idx_name] else ""
        if not name or name == "None":
            continue

        case_id = ""
        if idx_id is not None and row[idx_id]:
            case_id = str(row[idx_id]).strip()
        if not case_id:
            case_id = f"TC-{str(auto_id).zfill(3)}"
            auto_id += 1

        case_type = "Positive"
        if idx_type is not None and row[idx_type]:
            raw = str(row[idx_type]).strip()
            if any(k in raw for k in ["neg", "Neg", "부정", "negative", "Negative"]):
                case_type = "Negative"

        input_val = ""
        if idx_input is not None and row[idx_input]:
            input_val = str(row[idx_input]).strip()

        expected = ""
        if idx_expected is not None and row[idx_expected]:
            expected = str(row[idx_expected]).strip()

        category = ""
        if idx_category is not None and row[idx_category]:
            category = str(row[idx_category]).strip()
            if category:
                categories.add(category)

        cases.append({
            "id": case_id,
            "name": name,
            "type": case_type,
            "input": input_val,
            "expected": expected,
            "actual": "",
            "pf": "Pass",
            "owner": "",
            "date": "",
            "catName": category,
        })

    return {
        "cases": cases,
        "categories": list(categories),
        "headers": header,
        "total": len(cases),
    }


@router.post("/preview-headers")
async def preview_headers(file: UploadFile = File(...), sheet_name: str = Form("")):
    """엑셀 파일의 시트 목록과 헤더만 미리 반환"""
    contents = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"엑셀 파일 읽기 실패: {str(e)}")

    sheets = wb.sheetnames
    ws = wb[sheet_name] if sheet_name in sheets else wb.active
    first_row = next(ws.iter_rows(values_only=True), [])
    headers = [str(h).strip() if h else "" for h in first_row]

    return {"sheets": sheets, "headers": headers}
